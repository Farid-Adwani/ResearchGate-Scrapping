from parsel import Selector
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time
import requests
from openpyxl import load_workbook ,Workbook
import os
import shutil
from datetime import datetime


min_publications=0


def save_old_data():
    file_path = 'profile_data.xlsx'
    history_folder = 'history'

    if not os.path.exists(file_path):
        workbook =  Workbook()
        workbook.save(file_path)
        sheet = workbook.active
        headers = ['Image', 'Name','User-name', 'Affiliation', 'Reads', 'Citations', 'Co-authors-number', 'Co-authors', 'Cited-number', 'Cited', 'Publications Number', 'Publications', 'Skills']

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col_num)].width = 20
            workbook.save(file_path)
        print("Excel file created.")
    else:
        if not os.path.exists(history_folder):
            os.makedirs(history_folder)
        current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
        new_file_name = f"profile_data_{current_datetime}.xlsx"
        new_file_path = os.path.join(history_folder, new_file_name)
        shutil.copy(file_path, new_file_path)
        print("Data file copied to history folder with name ",new_file_name)

def scrape_profile_data(profile_link: str):
    global min_publications
    user_name= profile_link.split('/')[len(profile_link.split('/'))-1].split('?')[0]
    print(user_name)

    file_path = 'profile_data.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.active
    next_row = len(sheet['A']) + 1
    if user_name:
        for row in sheet.iter_rows(min_row=2, max_col=3):
            if row[2].value == user_name:
                print(f"A person with the name '{user_name}' already exists in the Excel file.")
                workbook.save(file_path)
                return []
    
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=True)
        page = browser.new_page()
        
        
        page.goto(profile_link,timeout=60000)
        profile_html = page.content()
        selector = Selector(text=profile_html)
        
        publications_number = selector.css('[data-testid="publicProfileStatsPublications"]::text').get()


        try:
            if int(publications_number)<min_publications:
                print(f"The publications number : {publications_number} is less than the minimum required number {min_publications}  !")
                return []
        except Exception as e:
            print(f"The publications number : {publications_number} is less than the minimum required number  {min_publications} !")
            return []
        

        reads = selector.css('[data-testid="publicProfileStatsReads"]::text').get()
        citations = selector.css('[data-testid="publicProfileStatsCitations"]::text').get()
        person_name = selector.css('.nova-legacy-e-text--size-xl ::text').get()
        profile_image = selector.css('.nova-legacy-e-avatar__img::attr(src)').get()
        affiliation = selector.css('.lite-page__side .gtm-institution-item .nova-legacy-v-entity-item__stack-item .nova-legacy-v-entity-item__title a::text').getall()
        
        

        publication_elements = page.query_selector_all('.nova-legacy-v-publication-item__title')
        publications = []
        for element in publication_elements:
            publications.append(element.inner_text())

        co_selector="" 
        if selector.css(('.coauthor__list')):
            co_selector=".nova-legacy-c-modal__body-content .coauthor__container a.nova-legacy-e-link"
            page.click('.coauthor__list')
            try:
                page.wait_for_selector(co_selector, timeout=10000)
            except TimeoutError:
                print("Timed out waiting for the co-author links to load.")
        else:
            co_selector="li.nova-legacy-e-list__item .nova-legacy-v-person-list-item__align-content a.nova-legacy-e-link"

        
        co_authors = []
        co_author_links = page.query_selector_all(co_selector)
        for link in co_author_links:
            href = 'https://www.researchgate.net/'+link.get_attribute('href')
            name = link.text_content()
            co_user_name= href.split('/')[len(href.split('/'))-1].split('?')[0]
            co_authors.append({'href': href, 'name': name,'user_name':co_user_name})

        page.keyboard.press('Escape')

        cited = []
        if selector.css(('a[data-lite*="cited"]')):
            page.click('a[data-lite*="cited"]')
            try:
                page.wait_for_selector('.nova-legacy-c-modal__body-content .network-cited-modal__container a.nova-legacy-e-link', timeout=10000)
            except TimeoutError:
                print("Timed out waiting for the Cited links to load.")
            cited_links = page.query_selector_all('.nova-legacy-c-modal__body-content .network-cited-modal__container a.nova-legacy-e-link')
            for link in cited_links:
                href = link.get_attribute('href')
                name = link.text_content()
                cited.append({'href': href, 'name': name})

        skills = selector.css('.js-target-skills a::text').getall()


        # Print the extracted data
        # print("Name:", person_name)
        # print("Image:", profile_image)
        # print("Affiliation:", affiliation)
        # print("Reads:", reads)
        # print("Citations:", citations)
        # print("Co-authors-number: ",len(co_authors))
        # print("Co-authors: ",co_authors)
        # print("Cited-number: ",len(cited))
        # print("Cited: ",cited)
        # print("Publications Number:", publications_number)
        # print("Publications:", publications)
        # print("skills: ",skills)


        

        # Download the image from the URL
        if profile_image:
                           
                    
            response = requests.get(profile_image)
            image_path = 'profile_image.jpg'
            with open(image_path, 'wb') as image_file:
                image_file.write(response.content)

            # Insert the image into the Excel sheet
            img = Image(image_path)
            img.width = 100
            img.height = 100
            sheet.column_dimensions[get_column_letter(1)].width = 15
            sheet.row_dimensions[next_row].height = 80
            sheet.add_image(img, f'A{next_row}')

            # Insert data into the Excel sheet
            if person_name:
                sheet.cell(row=next_row, column=2, value=person_name)
            if user_name:
                sheet.cell(row=next_row, column=3, value=user_name)
            if affiliation:
                sheet.cell(row=next_row, column=4, value=' | '.join(affiliation))
            if reads:
                sheet.cell(row=next_row, column=5, value=float(reads.replace(',', '')))
            if citations:
                sheet.cell(row=next_row, column=6, value=float(citations.replace(',', '')))
            if co_authors:
                sheet.cell(row=next_row, column=7, value=len(co_authors))
            if co_authors:
                sheet.cell(row=next_row, column=8, value=' | '.join([co_author['user_name'] for co_author in co_authors]))
            if cited:
                sheet.cell(row=next_row, column=9, value=len(cited))
            if cited:
                sheet.cell(row=next_row, column=10, value=' | '.join([cited_item['name'] for cited_item in cited]))
            if publications_number:
                sheet.cell(row=next_row, column=11, value=int(publications_number))
            if publications:
                sheet.cell(row=next_row, column=12, value=' | '.join(publications))
            if skills:
                sheet.cell(row=next_row, column=13, value=' | '.join(skills))

            
        else:   
            print("page dont exist :( :( :( :( :( :( :( :( :( :(")


        # Save the workbook as an Excel file
        workbook.save('profile_data.xlsx')


        browser.close()
        links = [author["href"] for author in co_authors]
        return links


nb=0
key=""
def search_profile(link):
    global nb
    print(" ")
    print(" ")
    print(" ")

    start=time.time()
    nb+=1
    result= []
    # link="https://www.researchgate.net/profile/Mohammad-Mirhoseini"
    print("Scrapping link : ",link)
      
    try:
        result=scrape_profile_data(link)
        if result:
            print("------------------------------------------------- link number ",nb," is done ------------------------------------")
    except Exception as e:
        print("")
        print(".................................................................")
        print("*********************** ERROOOOOR in link",nb," RETRYING ...")
        print(f"An error occurred while scraping the profile: {str(e)}")
        print(".................................................................")
        print("")
        
    end=time.time()
    elapsed_time = end - start
    print("Time Spent in iteration : ",elapsed_time," seconds")
    print(" ")
    print(" ")
    print(" ")
    return result


global_links=[]
key=input("Enter a link : ")
global_links.append(key)

save_old_data()

depth = -1
while depth<10:
    if not global_links:
        global_links=[]
    depth+=1
    print("-------------------------- DEPTH ",depth," --------------------------")
    for link in global_links :
        print(link)
    print("-------------------------- ",len(global_links)," LINKS --------------------------")

    temp_links=[]
    for link in global_links:
        res=search_profile(link)
        print(res)
        if not res:
            res=[]
        print(res)
        for r in res:
            temp_links.append(r)
        print(temp_links)

    global_links=temp_links
    print(global_links)


        

        
