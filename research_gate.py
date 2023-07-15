from parsel import Selector
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time
import requests
from openpyxl import load_workbook ,Workbook
import os
import shutil
from datetime import datetime


min_publications=0


def save_old_data():
    file_path = 'profile_data.xlsx'
    history_folder = 'history'

    if not os.path.exists(file_path):
        workbook =  Workbook()
        workbook.save(file_path)
        sheet = workbook.active
        headers = ['Image', 'Name', 'Affiliation', 'Reads', 'Citations', 'Co-authors-number', 'Co-authors', 'Cited-number', 'Cited', 'Publications Number', 'Publications', 'Skills']

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.column_dimensions[get_column_letter(col_num)].width = 20
            workbook.save(file_path)
        print("Excel file created.")
    else:
        if not os.path.exists(history_folder):
            os.makedirs(history_folder)
        current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
        new_file_name = f"profile_data_{current_datetime}.xlsx"
        new_file_path = os.path.join(history_folder, new_file_name)
        shutil.copy(file_path, new_file_path)
        print("Data file copied to history folder with name ",new_file_name)

def scrape_profile_data(profile_link: str):
    global min_publications
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=True)
        page = browser.new_page()
        
        
        page.goto(profile_link,timeout=60000)
        profile_html = page.content()
        selector = Selector(text=profile_html)
        
        publications_number = selector.css('[data-testid="publicProfileStatsPublications"]::text').get()


        try:
            if int(publications_number)<min_publications:
                print(f"The publications number : {publications_number} is less than the minimum required number {min_publications}  !")
                return False
        except Exception as e:
            print(f"The publications number : {publications_number} is less than the minimum required number  {min_publications} !")
            return False
        

        reads = selector.css('[data-testid="publicProfileStatsReads"]::text').get()
        citations = selector.css('[data-testid="publicProfileStatsCitations"]::text').get()
        person_name = selector.css('.nova-legacy-e-text--size-xl ::text').get()
        profile_image = selector.css('.nova-legacy-e-avatar__img::attr(src)').get()
        affiliation = selector.css('.lite-page__side .gtm-institution-item .nova-legacy-v-entity-item__stack-item .nova-legacy-v-entity-item__title a::text').getall()

        


        publication_elements = page.query_selector_all('.nova-legacy-v-publication-item__title')
        publications = []
        for element in publication_elements:
            publications.append(element.inner_text())

        co_selector="" 
        if selector.css(('.coauthor__list')):
            co_selector=".nova-legacy-c-modal__body-content .coauthor__container a.nova-legacy-e-link"
            page.click('.coauthor__list')
            try:
                page.wait_for_selector(co_selector, timeout=5000)
            except TimeoutError:
                print("Timed out waiting for the co-author links to load.")
        else:
            co_selector="li.nova-legacy-e-list__item .nova-legacy-v-person-list-item__align-content a.nova-legacy-e-link"

        
        co_authors = []
        co_author_links = page.query_selector_all(co_selector)
        for link in co_author_links:
            href = 'https://www.researchgate.net/'+link.get_attribute('href')
            name = link.text_content()
            co_authors.append({'href': href, 'name': name})

        page.keyboard.press('Escape')

        cited = []
        if selector.css(('a[data-lite*="cited"]')):
            page.click('a[data-lite*="cited"]')
            try:
                page.wait_for_selector('.nova-legacy-c-modal__body-content .network-cited-modal__container a.nova-legacy-e-link', timeout=5000)
            except TimeoutError:
                print("Timed out waiting for the Cited links to load.")
            cited_links = page.query_selector_all('.nova-legacy-c-modal__body-content .network-cited-modal__container a.nova-legacy-e-link')
            for link in cited_links:
                href = link.get_attribute('href')
                name = link.text_content()
                cited.append({'href': href, 'name': name})

        skills = selector.css('.js-target-skills a::text').getall()


        # Print the extracted data
        # print("Name:", person_name)
        # print("Image:", profile_image)
        # print("Affiliation:", affiliation)
        # print("Reads:", reads)
        # print("Citations:", citations)
        # print("Co-authors-number: ",len(co_authors))
        # print("Co-authors: ",co_authors)
        # print("Cited-number: ",len(cited))
        # print("Cited: ",cited)
        # print("Publications Number:", publications_number)
        # print("Publications:", publications)
        # print("skills: ",skills)


        file_path = 'profile_data.xlsx'
        workbook = load_workbook(file_path)
        sheet = workbook.active
        next_row = len(sheet['A']) + 1

        # Download the image from the URL
        if profile_image:
            for row in sheet.iter_rows(min_row=2, max_col=2):
                if row[1].value == person_name:
                    print(f"A person with the name '{person_name}' already exists in the Excel file.")
                    workbook.save(file_path)
                    browser.close()
                    return False
                
                    
            response = requests.get(profile_image)
            image_path = 'profile_image.jpg'
            with open(image_path, 'wb') as image_file:
                image_file.write(response.content)

            # Insert the image into the Excel sheet
            img = Image(image_path)
            img.width = 100
            img.height = 100
            sheet.column_dimensions[get_column_letter(1)].width = 15
            sheet.row_dimensions[next_row].height = 80
            sheet.add_image(img, f'A{next_row}')

            # Insert data into the Excel sheet
            if person_name:
                sheet.cell(row=next_row, column=2, value=person_name)
            if affiliation:
                sheet.cell(row=next_row, column=3, value=' | '.join(affiliation))
            if reads:
                sheet.cell(row=next_row, column=4, value=float(reads.replace(',', '')))
            if citations:
                sheet.cell(row=next_row, column=5, value=float(citations.replace(',', '')))
            if co_authors:
                sheet.cell(row=next_row, column=6, value=len(co_authors))
            if co_authors:
                sheet.cell(row=next_row, column=7, value=' | '.join([co_author['name'] for co_author in co_authors]))
            if cited:
                sheet.cell(row=next_row, column=8, value=len(cited))
            if cited:
                sheet.cell(row=next_row, column=9, value=' | '.join([cited_item['name'] for cited_item in cited]))
            if publications_number:
                sheet.cell(row=next_row, column=10, value=int(publications_number))
            if publications:
                sheet.cell(row=next_row, column=11, value=' | '.join(publications))
            if skills:
                sheet.cell(row=next_row, column=12, value=' | '.join(skills))

            
        else:   
            print("page dont exist :( :( :( :( :( :( :( :( :( :(")


        # Save the workbook as an Excel file
        workbook.save('profile_data.xlsx')


        browser.close()
        return True


def scrape_researchgate_publications(query: str,page_num):
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=True,timeout=5000)
        page = browser.new_page(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Firefox/101.0")
        links=[]
        authors = []
        next=False
        page.goto(f"https://www.researchgate.net/search/researcher?q={query}&page={page_num}")
        selector = Selector(text=page.content())
        authors=selector.css(".nova-legacy-c-card__body--spacing-inherit")
        for author in authors:
            if author.css('a[href*="publication/"]'):
                profile_url='https://www.researchgate.net/'+author.css('.nova-legacy-v-entity-item__title a::attr(href)').get().split('?')[0]
                links.append(profile_url)
        if selector.css(".nova-legacy-c-button-group__item:nth-child(9) a::attr(rel)").get():
            next=False
        else:
            next=True
        browser.close()
        return links,next






nb=0
page=0
next=True
key=""

key=input("Enter a searching key : ")
page=int(input("Enter the page to start with : "))-1
save_old_data()
while next:
    page+=1
    links,next=scrape_researchgate_publications(key,page)

    print("Scrapping ", len(links)," links from page ",page," ====> ")

    for link in links:
        print(" ")
        print(" ")
        print(" ")

        start=time.time()
        nb+=1
        # link="https://www.researchgate.net/profile/Mohammad-Mirhoseini"
        print("Scrapping link : ",link)
        
        try:
            result=scrape_profile_data(link)
            if result:
                print("------------------------------------------------- link number ",nb," is done ------------------------------------")
        except Exception as e:
            print("")
            print(".................................................................")
            print("*********************** ERROOOOOR in link",nb," RETRYING ...")
            print(f"An error occurred while scraping the profile: {str(e)}")
            print(".................................................................")
            print("")
        
        end=time.time()
        elapsed_time = end - start
        print("Time Spent in iteration : ",elapsed_time," seconds")
        print(" ")
        print(" ")
        print(" ")

        
